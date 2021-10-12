# -*- coding: utf-8 -*-
"""
Created on Thu Apr 29 16:38:15 2021

@author: jacob
"""

import numpy as np
import pandas as pd
from scipy.spatial import distance
import timeit
from openpyxl import load_workbook

start = timeit.default_timer()

#%% - import
# Importing data set at 20 Hz including:
# True time, heartbeat, speed, acceleration, X, Y, Name

df = pd.read_csv (r'C:\Users\jacob\Documents\Studie\Master\Afstudeerproject Papendal\Thesis\Python\CSVfiles\22-03-2021-jongens.csv', sep = ',', encoding = "ISO-8859-1")  # using encoding due to Loïs instead of Lois
MaxSpeedHR = pd.read_excel (r'C:\Users\jacob\Documents\Studie\Master\Afstudeerproject Papendal\Thesis\Python\MaxSpeedAndHR.xlsx', sep = ',', encoding = "ISO-8859-1")

pd.set_option('display.max_columns', 50)

df['Time'] = pd.to_datetime(df['Date/Time'])  # rename Date/Time to Time 
df.drop(['Date/Time'], axis=1, inplace=True)  # drop Date/Time
df['Date'] = pd.to_datetime(df['Time']).dt.date # panda to datetime

#%% calculate Timestamp
def transform (dtime):  # use function
    return dtime.timestamp()*1000
df['Timestamp'] = (df.apply(lambda row: transform(row.Time), axis = 1))  # calculate timestamp of every row

#%% - pre-processing data
df['X'].fillna(method='pad', inplace=True)  # remove 0 with previous row value
df['Y'].fillna(method='pad', inplace=True)  # remove 0 with previous row value

df = df.dropna() # remove rows with missing data

#%% calculate relative speed
df['Max Speed'] = df['Name'].map(MaxSpeedHR.set_index('Name')['Max Speed']) # create new column with individual max speed
df['relative speed'] = df['Speed'] / df['Max Speed'] * 100 
df.drop(['Max Speed'], axis=1, inplace=True)

#%% calculate relative HR
df['Max HR'] = df['Name'].map(MaxSpeedHR.set_index('Name')['Max HR']) # create new column with individual max HR
df['relative HR'] = df['heartbeat'] / df['Max HR'] * 100
df.drop(['Max HR'], axis=1, inplace=True)

#%% pre-process incorrect data
# remove data from players outside the field
df.loc[df['X'] > 20.5, ['Speed', 'relative speed', 'Acceleration', 'relative HR', 'heartbeat']] = 0
df.loc[df['X'] < -20.5, ['Speed', 'relative speed', 'Acceleration', 'relative HR', 'heartbeat']] = 0
df.loc[df['Y'] > 11.0, ['Speed', 'relative speed', 'Acceleration', 'relative HR', 'heartbeat']] = 0
df.loc[df['Y'] < -11.0, ['Speed', 'relative speed', 'Acceleration', 'relative HR', 'heartbeat']] = 0
df.loc[df['Speed'] > 8.0, ['Speed', 'relative speed']] = 0

# remove outliers
df.loc[df['Acceleration'] < -5.0, ['Acceleration', 'relative speed']] = 0
df.loc[df['Acceleration'] > 5.0, ['Acceleration', 'relative speed']] = 0

#%% - build list for every player 
df['Name'] = df['Name'].astype('category')
playerName = df['Name'].unique().tolist()
playerName = [x for x in playerName if str(x) != 'nan']  # to make sure there are no NaN's in the script

df_list = [df.loc[df['Name'] == name] for name in playerName]

#%% - calculate variables per player

# calculate Euclidean Distance 
def calc_Euclidean(X, Y):  # calculate distance covered
    return distance.euclidean(X, Y)
for p in df_list:
    p['Euclidean Distance'] = p.apply(lambda row: calc_Euclidean(row.X, row.Y), axis=1) 
    p['Euclidean Distance Diff'] = abs(p['Euclidean Distance'] - p['Euclidean Distance'].shift(1))
    p.loc[(p['Speed'] == 0) & (p['Acceleration'] == 0) , 'Euclidean Distance Diff'] = 0 # when speed & acceleration is 0, total distance should be 0
    p.loc[(p['Euclidean Distance Diff'] > 0.8), 'Euclidean Distance Diff'] = 0 # remove outliers
    #%% remove wrong distances due to leaving the field
    p['Diff Timestamp'] = p['Timestamp'] - p['Timestamp'].shift(1) 
    p.loc[(p['Diff Timestamp'] > 100.0),  'Euclidean Distance Diff'] = 0 # timestamp above 100.0 means they left the playing field
    p['Total Distance'] = np.cumsum(p['Euclidean Distance Diff']) # calculate total distance bij cumulative sum of all rows
    p.drop(['Diff Timestamp', 'Euclidean Distance', 'Diff Timestamp'], axis=1, inplace=True) # drop columns used by calculations
    #%% calculate training duration
    p['On Field'] = 1 # 
    p.loc[(p['X'] > 20.5) | (p['X'] < -20.5) | (p['Y'] > 11.0) | (p['Y'] < -11.0), 'On Field'] = 0  # if player is outside the field
    p['heartbeat'] = p['heartbeat'].fillna(0)  # change nan to 0
    p.loc[(p['Speed'] == 0) & (p['Acceleration'] == 0) & (p['heartbeat'] == 0), 'On Field'] = 0  # If there was no signal, speed & acceleration & heartbeat was 0
    p['On Field2'] = p.groupby(['Name'])['On Field'].cumsum() / 20  # calculate seconds on field
    p['Training Time in seconds'] = p['On Field2'] # make one column in seconds
    p['Training Time'] = pd.to_datetime(p["On Field2"], unit='s').dt.strftime("%H:%M:%S") # make one column in HH:MM:SS
    p.drop(['Time', 'On Field2'], axis=1, inplace=True)    
    #%% calculate meters per category relative speed
    p['70-80'] = p.loc[(p['relative speed'] >= 70) & (p['relative speed'] < 80), 'Euclidean Distance Diff']  # find distance covered when speed >= 70% and < 80%
    p['Meters 70-80%'] = p.groupby(['Name'])['70-80'].cumsum()  # cumulative sum of all meters between speed >= 70% and < 80%
    p['80-90'] = p.loc[(p['relative speed'] >= 80) & (p['relative speed'] < 90), 'Euclidean Distance Diff']  
    p['Meters 80-90%'] = p.groupby(['Name'])['80-90'].cumsum()  
    p['90-100'] = p.loc[(p['relative speed'] >= 90), 'Euclidean Distance Diff']  
    p['Meters 90-100%'] = p.groupby(['Name'])['90-100'].cumsum()
    p.drop(['70-80', '80-90', '90-100'], axis=1, inplace=True)  
    p['Meters 70-80%'] = p['Meters 70-80%'].fillna(value=0) # give all empty cells value of 0     
    p['Meters 80-90%'] = p['Meters 80-90%'].fillna(value=0)    
    p['Meters 90-100%'] = p['Meters 90-100%'].fillna(value=0)    
    #%% calculate HID 
    p['High Intensity Distance'] = p.loc[(p['relative speed'] >= 70), 'Euclidean Distance Diff']  # find distance covered when speed >+ 70%
    p.loc[(p['High Intensity Distance'] > 0.8) , 'High Intensity Distance'] = 0  # choose which value is preferred to exclude outliers
    p['HID'] = p.groupby(['Name'])['High Intensity Distance'].cumsum()  # cumulative sum of all HID meters
    p.drop(['Euclidean Distance Diff', 'High Intensity Distance'], axis=1, inplace=True) 
    p['HID'].fillna(0)
    p['HID'] = p['HID'].fillna(value=0)    
    #%% calculate Acc > 2.0 m/s2
    p['Acc > 2.0'] = 0 # add column with 0 as default value
    p['Acceleration'].replace(to_replace=0, method='ffill')
    p.loc[(p['Acceleration'] > 2.2) & (p['Speed'] > 2) , 'Acc > 2.0'] = 1 # find all rows that fulfills your conditions and set Acc > 2.0 m/s2 to 1
    p['Diff Acc'] = p['Acc > 2.0'] + p['Acc > 2.0'].shift(1)  # if 2 then acc was more or equal to 100 ms, if 0 or 1 then acc was less than 50 ms
    p['Diff Acc'] = p['Diff Acc']**2  # square the values to 0, 1, 4 to calculate differences
    p['Diff2 Acc'] = p['Diff Acc'] - p['Diff Acc'].shift(1)  # if 3 then acc condition is met
    p.loc[(p['Diff2 Acc'] == 3), 'Acc'] = 1  # find every unique acc
    p['Total Acc'] = np.cumsum(p['Acc'])  # count all acc
    p.drop(['Acc > 2.0', 'Acc', 'Diff Acc', 'Diff2 Acc'], axis=1, inplace=True)
    p['Total Acc'] = p['Total Acc'].fillna(value=0)    
    #%% calculate Decc < -2.0 m/s2
    p['Decc > 2.0'] = 0 # add a column with 0 as default value
    p.loc[(p['Acceleration'] < -2.2) & (p['Speed'] > 2) , 'Decc > 2.0'] = 1 
    p['Diff Decc'] = p['Decc > 2.0'] + p['Decc > 2.0'].shift(1)
    p['Diff Decc'] = p['Diff Decc']**2
    p['Diff2 Decc'] = p['Diff Decc'] - p['Diff Decc'].shift()
    p.loc[(p['Diff2 Decc'] == 3), 'Decc'] = 1
    p['Total Decc'] = np.cumsum(p['Decc'])
    p.drop(['Decc > 2.0', 'Decc', 'Diff Decc', 'Diff2 Decc'], axis=1, inplace=True)
    p['Total Decc'] = p['Total Decc'].fillna(value=0)    
    #%% calculate sprints 
    p['Sprints'] = 0 # add column with 0 as default value
    p.loc[(p['Speed'] > 4.75), 'Sprints'] = 1 # find all rows that fulfill conditions and set Sprints to 1
    p['Diff Sprints'] = p['Sprints'].rolling(window=20).sum()
    p.loc[(p['Diff Sprints'] == 20), 'Sprints2'] = 1  # find every sprint
    p['Sprints2'] = p['Sprints2'].fillna(value=0)    
    p['Diff2 Sprints'] = p['Sprints2'] - p['Sprints2'].shift(1)  # if 1 then sprint condition is met
    p.loc[(p['Diff2 Sprints'] == 1), '#Sprints'] = 1  # find all unique sprint
    p['#Sprints'] = p['#Sprints'].fillna(value=0)    
    p['Total Sprints'] = np.cumsum(p['#Sprints'])  # count all sprints
    p.drop(['Sprints', 'Diff Sprints', 'Sprints2', '#Sprints'], axis=1, inplace=True)
    #%% calculate Metabolic Power (MP) and Energy Expenditure (EE)
    # first 5 code lines are used to calculate MP
    p['alpha'] = np.degrees(np.arctan(9.81/p['Acceleration']))
    p['g'] = np.sqrt(p['Acceleration'] ** 2 + 9.81 ** 2)
    p['ES'] = np.tan(0.5*np.pi - np.arctan(9.81/p['Acceleration']))
    p['EM'] = p['g']/9.81
    p['EC'] = (155.4*p['ES'] ** 5 - 30.4*p['ES'] ** 4 - 43.3*p['ES'] ** 3 + 46.3*p['ES'] ** 2 + 19.5*p['ES'] + 3.6)*p['EM']
    p.loc[(p['On Field'] == 0), 'EC'] = 0  # EC is 0 when outside field
    p['MP'] = p['EC'] * p['Speed']
    p['MP Cumulative'] = np.cumsum(p['MP'])
    p['EE'] = p['MP Cumulative']/20000 # calculate Energy Expenditure per second (/20) and per kJ (/1000)
    p.drop(['alpha', 'g', 'ES', 'EM', 'EC'], axis=1, inplace=True)
    #%% calculate EE per MP zone (based on W/kg)
    p['EE 0-10'] = p.loc[(p['MP'] < 10), 'MP'] # select all EE when MP is between 10 W/kg
    p.loc[(p['On Field'] == 0), 'EE 0-10'] = 0 # outside field results in EE = 0
    p['EE 10-20'] = p.loc[(p['MP'] >= 10) & (p['MP'] < 20), 'MP']  
    p['EE 20-35'] = p.loc[(p['MP'] >= 20) & (p['MP'] < 35), 'MP'] 
    p['EE 35-55'] = p.loc[(p['MP'] >= 35) & (p['MP'] < 55), 'MP']  
    p['EE > 55'] = p.loc[(p['MP'] >= 55), 'MP']  
    p['EE zone 0-10'] = np.cumsum(p['EE 0-10'] / 20000) # count all EE 0-10
    p['EE zone 10-20'] = np.cumsum(p['EE 10-20'] / 20000)  
    p['EE zone 20-35'] = np.cumsum(p['EE 20-35'] / 20000)  
    p['EE zone 35-55'] = np.cumsum(p['EE 35-55'] / 20000)  
    p['EE zone > 55'] = np.cumsum(p['EE > 55'] / 20000)   
    p['EE zone 0-10'] = p['EE zone 0-10'].fillna(value=0) # give all empty cells value of 0 
    p['EE zone 10-20'] = p['EE zone 10-20'].fillna(value=0)    
    p['EE zone 20-35'] = p['EE zone 20-35'].fillna(value=0)    
    p['EE zone 35-55'] = p['EE zone 35-55'].fillna(value=0)    
    p['EE zone > 55'] = p['EE zone > 55'].fillna(value=0)    
df2 = pd.concat(df_list)  # make new Dataframe with variables calculated above

#%% - new Dataframe with calculated variables per player
TrainingVariables = pd.DataFrame(df2.groupby(['Name'], as_index = False)['Date','Training Time in seconds','Training Time', 'Total Distance', 'Meters 70-80%', 'Meters 80-90%', 'Meters 90-100%', 'HID', 'Total Acc', 'Total Decc', 'Total Sprints', 'EE zone 0-10', 'EE zone 10-20','EE zone 20-35', 'EE zone 35-55', 'EE zone > 55', 'EE'].max())

print(TrainingVariables)

# export TrainingVariables to Excel
TrainingVariables.to_excel('PythonOutputMales.xlsx', index = False)

#%% append training variables to Excel file

book = load_workbook('DatabaseMales.xlsx')
writer = pd.ExcelWriter('DatabaseMales.xlsx', engine='openpyxl') 
writer.book = book    

writer.sheets = {ws.title: ws for ws in book.worksheets}

for sheetname in writer.sheets:
    TrainingVariables.to_excel(writer,sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index = False,header= False)

writer.save()

stop = timeit.default_timer()
print('Time: ', stop - start)



